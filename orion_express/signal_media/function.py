import configparser
import csv
from datetime import datetime
import shutil
import json
import openpyxl
import random
from .path import *
import logging

LOG_FILENAME = 'loger.log'
logging.basicConfig(format='%(asctime)s-%(levelname)s-%(message)s',
                    filename=LOG_FILENAME, level=logging.INFO)


def open_file(dir):
    files = os.listdir(dir)
    return files


# Запись списка в файл
def write_list_to_file(data, file_name):
    try:
        with open(file_name, 'w') as file:
            for item in data:
                # file.write(str(item) + '\n')
                file.write(str(item) + '\n')
    except Exception as e:
        return e


# Чтение списка из файла
def read_list_from_file(file_name):
    try:
        with open(file_name, 'r') as file:
            lines = file.readlines()
        data_read = [eval(line.strip()) for line in lines]
        return data_read
    except Exception as e:
        return []


def write_blocks_to_file(blocks, file_path, old_name):
    new_file_name = old_name.replace(".xlsx", ".csv")
    new_file_name = new_file_name[:4] + "-" + new_file_name[4:6] + '-' + new_file_name[
                                                                         6:]  # Добавляем дефис между годом и месяцем
    csv_file_path = os.path.join(file_path, new_file_name)

    with open(csv_file_path, 'w', newline='', encoding='Windows-1251') as file:
        first_video = True
        for block in blocks:
            video_name = block[0]
            if not first_video and video_name != prev_video_name:
                file.write('\n')  # Записываем пустую строку между видео блоками
            file.write(';'.join(block) + ';\n')  # Объединяем элементы списка в строку с использованием точки с запятой
            prev_video_name = video_name
            first_video = False


def write_blocks_to_file_v2(blocks, file_path, old_name):
    csv_file_path = os.path.join(file_path, old_name)
    with open(csv_file_path, 'w', newline='') as file:
        first_video = True
        for block in blocks:
            video_name = block[0]
            if not first_video and video_name != prev_video_name:
                file.write('\n')  # Записываем пустую строку между видео блоками
            file.write(';'.join(block) + ';\n')  # Объединяем элементы списка в строку с использованием точки с запятой
            prev_video_name = video_name
            first_video = False


def find_green_ranges(filename):
    # Загрузка файла Excel
    workbook = openpyxl.load_workbook(filename)
    sheet = workbook.active
    # Создание списка для хранения диапазонов зеленых строк
    green_ranges = []
    start_row = None
    end_row = None
    # Перебор строк в файле Excel
    for row_index, row in enumerate(sheet.iter_rows(), start=1):
        if row[1].fill.start_color.index == 'FFB3FFC1':
            if start_row is None:
                start_row = row_index
        else:
            if start_row is not None:
                end_row = row_index
                # end_row = row_index - 1
                green_ranges.append([start_row, end_row])
                start_row = None
                end_row = None

    # Если последняя строка также зеленая, добавляем диапазон
    if start_row is not None:
        end_row = row_index
        green_ranges.append([start_row, end_row])
    return green_ranges


def read_settings(channel_name):
    config = configparser.ConfigParser()

    config.read(ini_path, encoding='UTF-8')

    sections_list = config.sections()

    spisok = config.get(channel_name, 'spisok')
    spisok_exception = config.get(channel_name, 'spisok_exception')

    spisok = [item.strip() for item in spisok.split(',')]
    spisok_exception = [item.strip() for item in spisok_exception.split(',')]

    return spisok, spisok_exception


def time_to_str(time):
    return time.strftime("%H:%M:%S")


def time_to_seconds(time_obj):
    if type(time_obj) == str:
        time_obj = datetime.strptime(time_obj, "%H:%M:%S").time()
    total_seconds = time_obj.hour * 3600 + time_obj.minute * 60 + time_obj.second
    return total_seconds


def seconds_to_time(seconds):
    hours = seconds // 3600
    minutes = (seconds % 3600) // 60
    seconds = seconds % 60
    time_format = f"{hours:02d}:{minutes:02d}:{seconds:02d}"
    return time_format


def block_to_time(dur, genre, block):
    # на вход идет макс.время блока, все видео которые можно добавить и жанр минус 1
    count = []
    for i in range(0, len(block)):
        if block[i][1] <= dur and block[i][2] == genre:
            count.append(i)
    temp = random.choice(count)
    return temp


def find_first_and_count(matrix, target):  # на вход список и число, вернет первое вхождение и список вхождений
    count = sum(1 for item in matrix if item[1] == target)
    first_index = next((index for index, item in enumerate(matrix) if item[1] == target), None)
    return first_index, count


def parser_v2_test(path_to_old, old_list, chanel_name):
    # input
    # return result_green

    spisok, spisok_exception = [], []
    info = ''
    spisok, spisok_exception = read_settings(chanel_name)
    result_green, temp_green, temp_green_without_simvol = [], [], []
    if os.name == 'nt':
        temp_path = path_to_old + '\\' + old_list
    elif os.name == 'posix':
        temp_path = path_to_old + '//' + old_list
    '''Открываем файл, проверяем двойным условием,
     должен быть зеленый+ выше должны быть ключевые слова vimb,кросс промо и.т.д.'''
    try:
        wb = openpyxl.reader.excel.load_workbook(filename=temp_path)
        sheets_list = wb.sheetnames
        sheet = wb[sheets_list[0]]
    except Exception as e:
        logging.info(f'Ошибка при обработке файла  Excel: {temp_path}')
        logging.info(e)
        logging.info(f"//////////Ошибка при обработке файла Excel: {temp_path}//////////////")
        info += (f"//////////Ошибка при обработке файла Excel: {temp_path}//////////////")
        print(f"Ошибка при обработке файла Excel: {e}")
        print(e)
        return None, info
    temp_green = find_green_ranges(temp_path)
    for i in range(0, len(temp_green)):
        if any(value in sheet.cell(row=temp_green[i][0] - 1, column=2).value for value in spisok_exception):
            logging.info(f'Ошибка при обработке файла  Excel: {path_to_old + old_list}'
                         f'\nВ строках {temp_green[i]}')
        elif any(value in sheet.cell(row=temp_green[i][0] - 1, column=2).value for value in spisok):
            temp_green_without_simvol.append(temp_green[i])
        else:
            print(f'Ошибка при обработке строк {temp_green[i]}')
            logging.info(f'Ошибка при обработке файла  Excel: {path_to_old + old_list}'
                         f'\nВ строках {temp_green[i]}')
            info += (f'Ошибка при обработке файла  Excel: {path_to_old + old_list}'
                     f'\nВ строках {temp_green[i]}')
    result_green = []
    for i in range(len(temp_green_without_simvol)):
        temp_row = []
        for j in range(len(temp_green_without_simvol[i])):
            cell_value = sheet.cell(row=temp_green_without_simvol[i][j], column=1).value
            temp_row.append(cell_value[:-3])
        result_green.append(temp_row)
    # Добавляем длительность

    durations = []
    # объеденяем блоки
    result_green_v2 = []
    i = 0
    while i < len(result_green):
        start = result_green[i][0]
        end = result_green[i][1]

        while i + 1 < len(result_green) and abs(time_to_seconds(result_green[i + 1][0]) - time_to_seconds(
                end)) <= 30:  # 180 Если разница между блоками составляент меньше 180 секунд, то надо объеденить блоки
            end = result_green[i + 1][1]
            i += 1

        result_green_v2.append([start, end])
        i += 1
    # конец костыля
    datetime_format = "%H:%M:%S"
    timestamps = [[datetime.strptime(time_str, datetime_format) for time_str in time_list] for time_list in
                  result_green_v2]
    for time_pair in timestamps:
        time_diff = time_pair[1] - time_pair[0]
        hours, remainder = divmod(time_diff.seconds, 3600)
        minutes, seconds = divmod(remainder, 60)
        duration_str = f"{hours:02}:{minutes:02}:{seconds:02}"
        durations.append(duration_str)
    for i in range(len(result_green_v2)):
        if i < len(durations):
            result_green_v2[i].append(durations[i])
    blocks_str = ''
    blocks_result = []
    # blocks_result = create_blocks_v4(video_file, result_green_v2, chanel_name)
    return result_green_v2, info


def create_blocks_v4_sarafan(path_to_video, result_green, chanel_name):  # out list
    # Если надо меньше роликов в блоке, в резинке
    # Открытие Excel-файла с роликами и их длительностью
    workbook = openpyxl.load_workbook(path_to_video)
    sheet = workbook.active
    sheet = workbook[chanel_name]
    result = ''
    videos, videos_rezinka, rand_temp_lis, rand_block_temp, rand_block_rezinka, rand_block_result, video_full = [], [], [], [], [], [], []
    # Извлечение роликов и их длительности из файла
    for row in sheet.iter_rows(min_row=1, values_only=True):
        video_name = row[0]
        duration = row[1]
        gorne = row[2]
        temp = row[3]
        if video_name and duration and duration != "None":
            if gorne == 8:
                videos_rezinka.append((video_name, time_to_seconds(duration), gorne, temp))
            else:
                videos.append((video_name, time_to_seconds(duration), gorne, temp))
    videos_sorted = sorted(videos, key=lambda x: x[1], reverse=True)
    videos_rezinka_sorted = sorted(videos_rezinka, key=lambda x: x[1], reverse=True)
    video_full = videos_sorted.copy()
    gorne_temp = 0
    gorne_temp_for_next = 0
    for i in range(0, len(result_green)):
        for l in range(1, 6):  # Если заканчиваются ролики с жанром, то он дублирует
            second_elements = [item[2] for item in videos_sorted]
            count_of_genre = second_elements.count(l)
            if count_of_genre == 0:
                for j in range(0, len(video_full)):
                    if video_full[j][2] == l:
                        videos_sorted.append(video_full[j])
                videos_sorted.sort(key=lambda x: x[1], reverse=True)
        del rand_block_temp[:]
        dur_temp = time_to_seconds(result_green[i][2])
        count = 0
        k = 0
        stopper = 0
        while dur_temp >= videos_sorted[-1][1] and stopper < 1000:
            gorne_temp = 0
            if dur_temp >= (videos_sorted[-1][1]):
                while len(videos_sorted) > 0:
                    cont_temp = 0
                    stopper += 1
                    if stopper > 100000:
                        break
                    if dur_temp < videos_sorted[-1][1]:
                        break
                    if videos_sorted[k][2] - 1 == gorne_temp and videos_sorted[k][1] <= dur_temp and (
                            dur_temp - videos_sorted[k][1]) >= 5:
                        gorne_temp = videos_sorted[k][2]
                        count_temp = block_to_time(dur_temp, gorne_temp, videos_sorted)
                        if videos_sorted[count_temp][2] == gorne_temp and \
                                videos_sorted[count_temp][1] <= dur_temp and (
                                dur_temp - videos_sorted[count_temp][1]) >= 5:
                            block = [

                                result_green[i][0],
                                videos_sorted[count_temp][0],
                                result_green[i][2],
                                seconds_to_time(videos_sorted[count_temp][1]),

                            ]
                            rand_block_temp.append(block)
                            dur_temp -= videos_sorted[count_temp][1]
                            videos_sorted.pop(count_temp)
                            count -= 1
                            k = 0
                            for l in range(1, 6):  # Если заканчиваются ролики с жанром, то он дублирует
                                second_elements = [item[2] for item in videos_sorted]
                                count_of_genre = second_elements.count(l)
                                if count_of_genre == 0:
                                    for j in range(0, len(video_full)):
                                        if video_full[j][2] == l:
                                            videos_sorted.append(video_full[j])
                                    videos_sorted.sort(key=lambda x: x[1], reverse=True)
                    elif k < len(videos_sorted) - 1:
                        k += 1  # kostil;
                    elif k == len(videos_sorted) - 1:
                        gorne_temp += 1
                        k = 0
                    if gorne_temp == 5:  # 5 жанров
                        gorne_temp = 0
                count += 1
            if len(videos_sorted) == 0:
                videos_sorted = video_full.copy()
            # count += 1
        if dur_temp > 13:  # исключаем вероятность того, кольцо будет больше 13-ти секунд

            count = 0
            while not 3 <= dur_temp <= 13:
                if 3 <= dur_temp - video_full[count][1] <= 13:
                    block = [
                        result_green[i][0],
                        video_full[count][0],
                        result_green[i][2],
                        seconds_to_time(video_full[count][1]),
                    ]
                    rand_block_temp.append(block)
                    dur_temp -= video_full[count][1]
                count += 1

        for j in range(0, len(rand_block_temp)):  # это рандомизация
            rand_block_result.append(rand_block_temp[j])
        else:
            if dur_temp == 0:
                continue
            else:
                count = len(videos_rezinka_sorted) - 1
                # иду от обротного, чтобы было больше роликов
                # look1
                # костыль для нст
                if len(videos_rezinka_sorted) == 1:
                    while dur_temp > 0:
                        block = [
                            result_green[i][0],
                            videos_rezinka_sorted[count][0],
                            result_green[i][2],
                            seconds_to_time(dur_temp),

                        ]
                        rand_block_result.append(block)
                        dur_temp -= videos_rezinka_sorted[count][1]
                # конец костыля
                while dur_temp > 0:  # nado horosho podumat
                    # Создаем новый блок с резиновой и добавляем его в список blocks
                    block = [
                        result_green[i][0],
                        videos_rezinka_sorted[count][0],
                        result_green[i][2],
                        seconds_to_time(dur_temp),

                    ]
                    rand_block_result.append(block)
                    dur_temp -= videos_rezinka_sorted[count][1]
                    count -= 1
                    # if count == len(videos_rezinka_sorted):
                    #     count = 0
                    if count == 0:
                        count = len(videos_rezinka_sorted) - 1  # reload
    return rand_block_result


def create_blocks_v4_rand(path_to_video, result_green, chanel_name):  # out list
    # Если надо меньше роликов в блоке, в резинке
    # Открытие Excel-файла с роликами и их длительностью
    workbook = openpyxl.load_workbook(path_to_video)
    sheet = workbook.active
    sheet = workbook[chanel_name]
    videos, videos_rezinka, rand_temp_lis, rand_block_temp, rand_block_rezinka, rand_block_result, video_full = [], [], [], [], [], [], []
    # Извлечение роликов и их длительности из файла
    for row in sheet.iter_rows(min_row=1, values_only=True):
        video_name = row[0]
        duration = row[1]
        if video_name and duration and duration != "None":
            if 'резиновая' in video_name.lower():
                videos_rezinka.append((video_name, time_to_seconds(duration)))
            else:
                videos.append((video_name, time_to_seconds(duration)))
    videos_sorted = sorted(videos, key=lambda x: x[1], reverse=True)
    videos_rezinka_sorted = sorted(videos_rezinka, key=lambda x: x[1], reverse=True)
    video_full = videos_sorted.copy()
    video_result = []
    for i in range(0, len(result_green)):  # start 0
        del rand_block_temp[:]
        dur_temp = time_to_seconds(result_green[i][2])
        count = 0
        while dur_temp >= videos_sorted[-1][1]:
            if count == len(videos_sorted):
                break
            if dur_temp >= (videos_sorted[count][
                1]):
                block = [
                    result_green[i][0],
                    videos_sorted[count][0],
                    result_green[i][2],
                    seconds_to_time(videos_sorted[count][1])
                ]
                rand_block_temp.append(block)
                dur_temp -= videos_sorted[count][1]
                videos_sorted.pop(count)
                count -= 1
            if len(videos_sorted) == 0:
                videos_sorted = video_full.copy()
            count += 1
        random.shuffle(rand_block_temp)
        for k in range(0, len(rand_block_temp)):  # это рандомизация
            rand_block_result.append(rand_block_temp[k])
        else:
            if dur_temp == 0:
                continue
            else:
                count = len(videos_rezinka_sorted) - 1
                # иду от обротного, чтобы было больше роликов
                # look1
                # костыль для нст
                if len(videos_rezinka_sorted) == 1:
                    while dur_temp > 0:
                        block = [
                            result_green[i][0],
                            videos_rezinka_sorted[count][0],
                            result_green[i][2],
                            seconds_to_time(videos_rezinka_sorted[count][1])
                        ]
                        rand_block_result.append(block)
                        dur_temp -= videos_rezinka_sorted[count][1]
                # конец костыля
                while dur_temp > 0:  # nado horosho podumat
                    # Создаем новый блок с резиновой и добавляем его в список blocks
                    block = [
                        result_green[i][0],
                        videos_rezinka_sorted[count][0],
                        result_green[i][2],
                        seconds_to_time(videos_rezinka_sorted[count][1])
                    ]
                    rand_block_result.append(block)
                    dur_temp -= videos_rezinka_sorted[count][1]
                    count -= 1
                    # if count == len(videos_rezinka_sorted):
                    #     count = 0
                    if count == 0:
                        count = len(videos_rezinka_sorted) - 1  # reload
    return rand_block_result


def plst_to_log():
    # TODO подумать куда вставить, доделать

    src = [path_before, path_after_rename, path_to_out, path_after_kzpl, dir_for_blocks]
    trg = path_to_log

    try:
        count = 0

        for i in range(len(src)):
            files = os.listdir(src[i])
            print(files)
            for fname in files:
                current_time = datetime.now().strftime("%Y%m%d_%H%M%S")
                shutil.copy2(os.path.join(src[i], fname), os.path.join(trg, current_time + fname))
                count += 1

            shutil.rmtree(src[i])
            os.mkdir(src[i])

        result = {'error': False, 'message': f'Был очищен список, там было {count} файлов'}
    except Exception as e:
        result = {'error': True, 'message': str(e)}

    return result


def rename_recopy(file_name, year):  # функция переимонования из исходного для такого, что подходит кзплу # работает со
    # всеми каналами, но название должно быть такого вида:
    # Путевка Сарафан (Орион) 17.07.22.xlsx
    # Путевка (Орион) МУЛЬТ HD 17.07.2022.xlsx
    # Путевка (Орион) Синема на 18.07.2022.xlsx
    # Путевка (Орион) НСТ на 18.07.2022.xlsx
    # Наука 18.07.xlsx
    # МП 17.07.xlsx
    # Т24 17.07.xlsx
    os.makedirs(os.path.dirname(path_after_rename), exist_ok=True)
    os.makedirs(os.path.dirname(path_after_kzpl), exist_ok=True)
    os.makedirs(os.path.dirname(dir_for_blocks), exist_ok=True)
    os.makedirs(os.path.dirname(path_to_log), exist_ok=True)
    temp_name_file = ' '.join(file_name)
    temp = file_name[-1].split('.')  # тут хранится дата(индекс от 0 до 2-х и 3-ый индекс-расширение
    if len(temp[2]) == 2:
        temp[2] = '20' + temp[2]
    # print(path_before, file_name)
    if file_name[0] == 'Путевка' and file_name[1].lower() == 'сарафан'.lower():  # проверка, с той путевкой мы работаем
        temp = temp[2], temp[1], temp[0], '_', file_name[1], '.', temp[3]
        res = ''.join(temp)
        shutil.copyfile(path_before + temp_name_file,
                        path_after_rename + res)  # берет файл из одной папки и копирует в ту же
    elif file_name[0] == 'Путевка' and file_name[2].lower() == 'МУЛЬТ'.lower():
        temp = temp[2], temp[1], temp[0], '_', file_name[2].capitalize(), '.', temp[3]
        res = ''.join(temp)
        shutil.copyfile(path_before + temp_name_file,
                        path_after_rename + res)  # берет файл из одной папки и копирует в ту же
    elif file_name[0] == 'Путевка' and file_name[2].lower() == 'нст':
        temp = temp[2], temp[1], temp[0], '_', file_name[2].capitalize(), '.', temp[3]
        res = ''.join(temp)
        shutil.copyfile(path_before + temp_name_file,
                        path_after_rename + res)  # берет файл из одной папки и копирует в ту же
    elif file_name[0] == 'Путевка' and file_name[2].lower() == 'синема':
        temp = temp[2], temp[1], temp[0], '_', 'CINEMA', '.', temp[3]
        res = ''.join(temp)
        shutil.copyfile(path_before + temp_name_file,
                        path_after_rename + res)  # берет файл из одной папки и копирует в ту же
    elif file_name[0].lower() == 'наука':  # проверка, с той ли путевкой  мы работаем
        temp = str(year), temp[1], temp[0], '_', file_name[0], '.', temp[2]
        res = ''.join(temp)
        shutil.copyfile(path_before + temp_name_file,
                        path_after_rename + res)  # берет файл из одной папки и копирует в ту же
    elif file_name[0].lower() == 'Т24'.lower():  # проверка, с той ли путевкой  мы работаем
        temp = str(year), temp[1], temp[0], '_', file_name[0], '.', temp[2]
        res = ''.join(temp)
        shutil.copyfile(path_before + temp_name_file,
                        path_after_rename + res)  # берет файл из одной папки и копирует в ту же
    elif file_name[0].lower() == 'МП'.lower():  # проверка, с той ли путевкой  мы работаем
        temp = str(year), temp[1], temp[0], '_', file_name[0], '.', temp[2]
        res = ''.join(temp)
        shutil.copyfile(path_before + temp_name_file,
                        path_after_rename + res)  # берет файл из одной папки и копирует в ту же


def rename_recopy_v2(file_name, year):
    # функция переимонования из исходного для такоого, что подходит кзплу # работает со
    # всеми каналами, но название должно быть такого вида:
    # Путевка Сарафан (Орион) 17.07.22.xlsx
    # Путевка (Орион) МУЛЬТ HD 17.07.2022.xlsx
    # Путевка (Орион) Синема на 18.07.2022.xlsx
    # Путевка (Орион) НСТ на 18.07.2022.xlsx
    # Наука 18.07.xlsx
    # МП 17.07.xlsx
    # Т24 17.07.xlsx

    if file_name[len(file_name) - 1] == 'xls':
        print('xyeta ')

    temp_name_file = '_'.join(file_name)
    temp = file_name[-1].split('.')  # тут хранится дата(индекс от 0 до 2-х и 3-ый индекс-расширение
    if len(temp[2]) == 2:
        temp[2] = '20' + temp[2]
    if file_name[0] == 'Путевка' and file_name[1].lower() == 'сарафан'.lower():  # проверка, с той путевкой мы работаем
        print('asd')
        temp = temp[2], temp[1], temp[0], '_', file_name[1], '.', temp[3]
        res = ''.join(temp)
        shutil.copyfile(path_before + temp_name_file,
                        path_after_rename + res)  # берет файл из одной папки и копирует в ту же
    elif file_name[0] == 'Путевка' and file_name[2].lower() == 'МУЛЬТ'.lower():
        temp = temp[2], temp[1], temp[0], '_', file_name[2].capitalize(), '.', temp[3]
        res = ''.join(temp)
        shutil.copyfile(path_before + temp_name_file,
                        path_after_rename + res)  # берет файл из одной папки и копирует в ту же
    elif file_name[0] == 'Путевка' and file_name[2].lower() == 'нст':
        temp = temp[2], temp[1], temp[0], '_', file_name[2].capitalize(), '.', temp[3]
        res = ''.join(temp)
        shutil.copyfile(path_before + temp_name_file,
                        path_after_rename + res)  # берет файл из одной папки и копирует в ту же
    elif file_name[0] == 'Путевка' and file_name[2].lower() == 'синема':
        temp = temp[2], temp[1], temp[0], '_', 'CINEMA', '.', temp[3]
        res = ''.join(temp)
        shutil.copyfile(path_before + temp_name_file,
                        path_after_rename + res)  # берет файл из одной папки и копирует в ту же
    elif file_name[0].lower() == 'наука':  # проверка, с той ли путевкой  мы работаем
        temp = str(year), temp[1], temp[0], '_', file_name[0], '.', temp[2]
        res = ''.join(temp)
        shutil.copyfile(path_before + temp_name_file,
                        path_after_rename + res)  # берет файл из одной папки и копирует в ту же
    elif file_name[0].lower() == 'Т24'.lower():  # проверка, с той ли путевкой  мы работаем
        temp = str(year), temp[1], temp[0], '_', file_name[0], '.', temp[2]
        res = ''.join(temp)
        shutil.copyfile(path_before + temp_name_file,
                        path_after_rename + res)  # берет файл из одной папки и копирует в ту же
    elif file_name[0].lower() == 'МП'.lower():  # проверка, с той ли путевкой  мы работаем
        temp = str(year), temp[1], temp[0], '_', file_name[0], '.', temp[2]
        res = ''.join(temp)
        shutil.copyfile(path_before + temp_name_file,
                        path_after_rename + res)  # берет файл из одной папки и копирует в ту же


'''Работа с файлами:
Копирование на сервера, проверка на корректность, и.т.д.'''


def open_file(dir):
    files = os.listdir(dir)
    return files


def copy_after_kzpl(file_name):  # на плейлисты из папки out, раскидываем по резервным серверам
    # s=file_name.split('.')
    if 'мп'.lower() in file_name[0].lower():  # проверка, с той путевкой мы работаем
        shutil.copyfile(path_after_kzpl + file_name[0],
                        hd_out_osn + 'MyPlanetHD/' + file_name[
                            0])  # берет файл из одной папки и копирует в сервер #hd osn
        shutil.copyfile(path_after_kzpl + file_name[0],
                        hd_out_reserve + 'MyPlanetHD/' + file_name[0])
    elif 'сарафан'.lower() in file_name[0].lower():  # проверка, с той путевкой мы работаем
        shutil.copyfile(path_after_kzpl + file_name[0],  # берет файл из одной папки и копирует в сервер #hd osn
                        sd_out_osn + 'Сарафан/' + file_name[0])
        shutil.copyfile(path_after_kzpl + file_name[0],
                        sd_out_reserve + 'Сарафан/' + file_name[0])
    elif 'наука'.lower() in file_name[0].lower():  # проверка, с той путевкой мы работаем
        shutil.copyfile(path_after_kzpl + file_name[0],
                        hd_out_osn + 'NaukaHD/' + file_name[0])  # берет файл из одной папки и копирует в сервер #hd osn
        shutil.copyfile(path_after_kzpl + file_name[0],
                        hd_out_reserve + 'NaukaHD/' + file_name[0])
    elif 'мульт'.lower() in file_name[0].lower():  # проверка, с той путевкой мы работаем
        shutil.copyfile(path_after_kzpl + file_name[0],
                        hd_out_osn + 'MultHD/' + file_name[0])  # берет файл из одной папки и копирует в сервер #hd osn
        shutil.copyfile(path_after_kzpl + file_name[0],
                        hd_out_reserve + 'MultHD/' + file_name[0])
    elif 'нст'.lower() in file_name[0].lower():  # проверка, с той путевкой мы работаем
        shutil.copyfile(path_after_kzpl + file_name[0],
                        sd_out_osn + 'NST/' + file_name[0])
        shutil.copyfile(path_after_kzpl + file_name[0],
                        sd_out_reserve + 'NST/' + file_name[0])  # берет файл из одной папки и копирует в сервер #hd osn
    elif 'cinema'.lower() in file_name[0].lower():  # проверка, с той путевкой мы работаем
        shutil.copyfile(path_after_kzpl + file_name[0],
                        sd_out_osn + 'Cinema/' + file_name[0])  # берет файл из одной папки и копирует в сервер #hd osn
        shutil.copyfile(path_after_kzpl + file_name[0],
                        sd_out_reserve + 'Cinema/' + file_name[0])
    elif 'т24'.lower() in file_name[0].lower():  # проверка, с той путевкой мы работаем
        shutil.copyfile(path_after_kzpl + file_name[0],
                        sd_out_osn + 'Techno24/' + file_name[
                            0])  # берет файл из одной папки и копирует в сервер #hd osn
        shutil.copyfile(path_after_kzpl + file_name[0],
                        sd_out_reserve + 'Techno24/' + file_name[0])


def copy_after_kzpl_all(file_name):  # на плейлисты из папки out, раскидываем по серверам
    # s=file_name.split('.')
    if 'мп'.lower() in file_name[0].lower():  # проверка, с той путевкой мы работаем
        shutil.copyfile(path_after_kzpl + file_name[0],
                        hd_out_osn + 'MyPlanetHD/' + file_name[
                            0])  # берет файл из одной папки и копирует в сервер #hd osn
        shutil.copyfile(path_after_kzpl + file_name[0],
                        hd_out_reserve + 'MyPlanetHD/' + file_name[0])
    elif 'сарафан'.lower() in file_name[0].lower():  # проверка, с той путевкой мы работаем
        shutil.copyfile(path_after_kzpl + file_name[0],  # берет файл из одной папки и копирует в сервер #hd osn
                        sd_out_osn + 'Сарафан/' + file_name[0])
        shutil.copyfile(path_after_kzpl + file_name[0],
                        sd_out_reserve + 'Сарафан/' + file_name[0])
    elif 'наука'.lower() in file_name[0].lower():  # проверка, с той путевкой мы работаем
        shutil.copyfile(path_after_kzpl + file_name[0],
                        hd_out_osn + 'NaukaHD/' + file_name[0])  # берет файл из одной папки и копирует в сервер #hd osn
        shutil.copyfile(path_after_kzpl + file_name[0],
                        hd_out_reserve + 'NaukaHD/' + file_name[0])
    elif 'мульт'.lower() in file_name[0].lower():  # проверка, с той путевкой мы работаем
        shutil.copyfile(path_after_kzpl + file_name[0],
                        hd_out_osn + 'MultHD/' + file_name[0])  # берет файл из одной папки и копирует в сервер #hd osn
        shutil.copyfile(path_after_kzpl + file_name[0],
                        hd_out_reserve + 'MultHD/' + file_name[0])
    elif 'нст'.lower() in file_name[0].lower():  # проверка, с той путевкой мы работаем
        shutil.copyfile(path_after_kzpl + file_name[0],
                        sd_out_osn + 'NST/' + file_name[0])
        shutil.copyfile(path_after_kzpl + file_name[0],
                        sd_out_reserve + 'NST/' + file_name[0])  # берет файл из одной папки и копирует в сервер #hd osn
    elif 'cinema'.lower() in file_name[0].lower():  # проверка, с той путевкой мы работаем
        shutil.copyfile(path_after_kzpl + file_name[0],
                        sd_out_osn + 'Cinema/' + file_name[0])  # берет файл из одной папки и копирует в сервер #hd osn
        shutil.copyfile(path_after_kzpl + file_name[0],
                        sd_out_reserve + 'Cinema/' + file_name[0])
    elif 'т24'.lower() in file_name[0].lower():  # проверка, с той путевкой мы работаем
        shutil.copyfile(path_after_kzpl + file_name[0],
                        sd_out_osn + 'Techno24/' + file_name[
                            0])  # берет файл из одной папки и копирует в сервер #hd osn
        shutil.copyfile(path_after_kzpl + file_name[0],
                        sd_out_reserve + 'Techno24/' + file_name[0])


def sort_to_check(plst_to_sort):  # на плейлисты из папки out, раскидываем по серверам
    # s=file_name.split('.')
    new_plst = []
    count = 0
    for i in range(0, len(plst_to_sort)):
        if 'мп'.lower() in plst_to_sort[i].lower():
            new_plst.append(plst_to_sort[i])
    for i in range(0, len(plst_to_sort)):
        if 'сарафан'.lower() in plst_to_sort[i].lower():
            new_plst.append(plst_to_sort[i])
    for i in range(0, len(plst_to_sort)):
        if 'наука'.lower() in plst_to_sort[i].lower():
            new_plst.append(plst_to_sort[i])
    for i in range(0, len(plst_to_sort)):
        if 'мульт'.lower() in plst_to_sort[i].lower():
            new_plst.append(plst_to_sort[i])
    for i in range(0, len(plst_to_sort)):
        if 'нст'.lower() in plst_to_sort[i].lower():
            new_plst.append(plst_to_sort[i])
    for i in range(0, len(plst_to_sort)):
        if 'cinema'.lower() in plst_to_sort[i].lower():
            new_plst.append(plst_to_sort[i])
    for i in range(0, len(plst_to_sort)):
        if 'т24'.lower() in plst_to_sort[i].lower():
            new_plst.append(plst_to_sort[i])
    return plst_to_sort


'''
Тестирование плейлистов
'''


def test_file_size():
    file_sizes = {}
    result = ''
    items = os.listdir(dir_for_blocks)
    for item in items:
        item_path = os.path.join(dir_for_blocks, item)
        if os.path.isfile(item_path):
            file_size = os.path.getsize(item_path)
            file_sizes[item] = file_size
    for file, size in file_sizes.items():
        if size < 8 * 1024:
            print((f'Маленький файл, проверь: {file}, Размер: {size} байт'))
            result += ((f'Маленький файл, проверь: {file}, Размер: {size} байт'))
    return file_sizes


def check_sec(path, filename, path_to_old, old_list):  # todo доделать нормальный вывод в json
    result = ''  # Результат, который мы будем возвращать
    liner, g, timer = [], [], []
    temp = []
    error_time = ''  # Строка для записи времени ошибок
    try:
        with open(path + filename, newline='') as file:
            playlist = csv.reader(file, delimiter=';')
            for row in playlist:
                if len(row) != 0:
                    liner.append(row)
    except Exception as e:
        print(e)
    for i in range(0, len(liner)):
        if len(liner[i]) != 0:
            temp = datetime.strptime(liner[i][0], '%H:%M:%S')
            timer.append(temp.minute * 60 + temp.hour * 60 * 60 + temp.second)
        else:
            timer.append('')
    temp2 = [[] for j in range(0, len(liner))]
    try:  # заполняем новый массив секундами от начала старта врезки и до конца
        for i in range(0, len(liner)):
            if len(liner[i]) != 0:
                temp = datetime.strptime(liner[i][0], '%H:%M:%S')
                temp2[i].append(temp.minute * 60 + temp.hour * 60 * 60 + temp.second)
                temp = datetime.strptime(liner[i][2], '%H:%M:%S')
                temp2[i].append(temp2[i][0] + temp.minute * 60 + temp.hour * 60 * 60 + temp.second)
    except Exception as e:
        print(e)  # ToDO переделать вывод в json
    result = ''
    # работа с плейлистом до кзпл-а
    timer_after_kzpl = []
    j = 0
    try:
        for i in range(0, len(temp2)):
            if temp2[i - 1] != temp2[i]:
                timer_after_kzpl.append(temp2[i])
    except Exception as e:
        print(e)  # ToDO переделать вывод в json
    try:
        wb = openpyxl.reader.excel.load_workbook(filename=path_to_old + old_list)
        sheets_list = wb.sheetnames
        sheet = wb[sheets_list[0]]
        for i in range(1, sheet.max_row):
            green = sheet.cell(row=i, column=1).fill.start_color.index  # FFB3FFC1 зеленый цвет
            if green == 'FFB3FFC1':
                green_value = sheet.cell(row=i, column=1).value[:-3]  # убираем кадры
                green_value = datetime.strptime(green_value, '%H:%M:%S')
                green_value_sec = green_value.minute * 60 + green_value.hour * 60 * 60 + green_value.second

                temper = 5  # этот костыль надо переделать
                for j in range(0, len(timer_after_kzpl)):
                    if timer_after_kzpl[j][0] <= green_value_sec <= timer_after_kzpl[j][1]:
                        temper = 0
                        break
                if temper == 5:
                    error_time += str(green_value)[10:] + '\n'
    except Exception as e:
        print(e)  # ToDO переделать вывод в json
    print(result)
    if error_time != '':
        result = {'error': True, 'message': 'В плейлисте ' + old_list + " есть ошибки на:", 'error_time': error_time}
    else:
        result = {'error': False, 'message': 'В плейлисте ' + old_list + " ошибок нет"}

    return result


def check_sec_to_min(path, filename):
    result = ''  # Результат, который мы будем возвращать
    liner, g, timer = [], [], []
    with open(path + filename, newline='') as file:
        playlist = csv.reader(file, delimiter=';')
        for row in playlist:
            liner.append(row)

    for i in range(0, len(liner)):
        if len(liner[i]) != 0:
            temp = datetime.strptime(liner[i][0], '%H:%M:%S')
            timer.append(temp.minute * 60 + temp.hour * 60 * 60 + temp.second)
        else:
            timer.append('')

    for i in range(0, len(timer) - 4):
        if type(timer[i + 1]) == str:
            if 0 < timer[i + 2] - timer[i] < 180:
                if result == '':
                    result += 'В плейлисте:' + filename + ' проверь строки ' + str(i + 1) + '-' + str(i + 3) + ';'

    if len(result) < 3:
        result = {'error': False, 'message': 'В плейлисте:' + filename + ' ошибок нет'}
    else:
        result = {'error': True, 'message': result}

    return result


if __name__ == '__main__':
    pass
